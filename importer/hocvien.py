from openpyxl import load_workbook
import re

IMPORT_FILE = "Danh sach nv nddh2019.xlsx"

wbs = load_workbook(IMPORT_FILE)
print(wbs.sheetnames)

# Find start of quiz table
start_of_quiz_table = False
user_list = []
for row in wbs.worksheets[0]:
    if start_of_quiz_table and row[1].value and row[2].value and row[3].value and row[4].value:
        user_list.append({"id": row[0].value, "fullname": row[1].value, "dob": row[2].value, "sex": row[3].value, "dept": row[4].value})
    else:
        print(row[0].value)
    if row[0].value == "mahocvien":
        start_of_quiz_table = True
        print("Start of quiz table found, starting to import")
print(user_list)