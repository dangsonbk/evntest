from openpyxl import load_workbook
import re

IMPORT_FILE = "Bo de thi giu bac khu khoang VH1.xlsx"

wbs = load_workbook(IMPORT_FILE)
print(wbs.sheetnames)

for wb in wbs:
    # Find start of quiz table
    start_of_quiz_table = False
    question_list = {"name": wb.title, "questions": []}
    for row in wb:
        if start_of_quiz_table and row[0].value:
            regx = re.search("(Câu \d+).(.+)", row[1].value)
            if regx:
                multichoices = {}
                question = regx.group(2).strip()
                if question[-1] == ":" or question[-1] == "?":
                    question = question[:-1]
                for q in [row[2].value, row[3].value, row[4].value, row[5].value]:
                    if q is None:
                        print("Wrong format, bypass this option")
                        continue
                    regx = re.search("([abcd])\.(.+)", q)
                    if regx:
                        multichoices[regx.group(1).strip().upper()] = regx.group(2).strip()
                question_list["questions"].append({"question": question, "multichoices": multichoices, "answer": row[6].value})
            else:
                print("Question format does not match: ", row[1].value)
        if row[0].value == "STT":
            start_of_quiz_table = True
            print("Start of quiz table found, starting to import", wb.title)
    print(question_list)